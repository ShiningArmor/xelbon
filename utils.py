import xlrd
from openpyxl.workbook import Workbook
from openpyxl.workbook import Workbook as openpyxlWorkbook

from openpyxl.reader.excel import load_workbook, InvalidFileException

def xls2xlsx(filename):
    # first open using xlrd

    xlsBook = xlrd.open_workbook(filename)
    workbook = openpyxlWorkbook()

    for i in xrange(0, xlsBook.nsheets):
        xlsSheet = xlsBook.sheet_by_index(i)
        sheet = workbook.active if i == 0 else workbook.create_sheet()
        sheet.title = xlsSheet.name

        for row in xrange(0, xlsSheet.nrows):
            for col in xrange(0, xlsSheet.ncols):
                sheet.cell(row=row + 1, column=col + 1).value = xlsSheet.cell_value(row, col)

    return workbook


def letter2number(values):
    letters = '0ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    result = 0
    for value in values:
        result *= 26
        result += letters.index(value)
    return result

def index2excel(row, col):
    letters = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    result = []
    while col:
        col, rem = divmod(col-1, 26)
        result[:0] = letters[rem]
    return ''.join(result) + str(row)